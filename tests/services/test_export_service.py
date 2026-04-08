"""
Integration tests for zeiterfassung/services/export_service.py (MAJ-003 fix).

Uses shared db_conn fixture from conftest.py.
Injects a real EntryService instance with pre-inserted entries into ExportService.
"""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import pytest

from zeiterfassung.config import Settings
from zeiterfassung.domain.models import EntryType, TimeEntry
from zeiterfassung.repository.entry_repo import EntryRepository
from zeiterfassung.services.entry_service import EntryService
from zeiterfassung.services.export_service import ExportService

_NOW = datetime.datetime(2026, 4, 1, 9, 0, 0)
_TARGET = 480


def _setup_services(db_conn) -> tuple[EntryService, ExportService]:
    """Create EntryService and ExportService backed by an in-memory DB."""
    settings = Settings(weekly_hours=40.0, bundesland="BY")
    repo = EntryRepository(db_conn)
    entry_service = EntryService(repo, settings)
    export_service = ExportService(entry_service)
    return entry_service, export_service


class TestExportExcel:
    """Tests for ExportService.export_excel (TEST-019 through TEST-021)."""

    def test_correct_headers_in_row_1(self, db_conn, tmp_path):
        """
        Export creates a file with correct headers in row 1 (TEST-019).

        Headers: Date, Type, Start, End, Pause (h), Delta (h), Running Saldo (h)
        """
        entry_service, export_service = _setup_services(db_conn)

        # Insert one entry
        entry_service.add_entry(
            date=datetime.date(2026, 4, 14),
            entry_type=EntryType.krank,
        )

        out = tmp_path / "test_export.xlsx"
        export_service.export_excel(
            from_date=datetime.date(2026, 4, 14),
            to_date=datetime.date(2026, 4, 14),
            output_path=out,
        )

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "Date", "Type", "Start", "End", "Pause (h)", "Delta (h)", "Running Saldo (h)"
        ]

    def test_empty_range_produces_only_header_and_summary(self, db_conn, tmp_path):
        """
        Export over an empty date range (no workdays, no entries) produces
        only the header row and summary row — no data rows.
        """
        _, export_service = _setup_services(db_conn)

        # Apr 11-12 is Saturday/Sunday — no workdays, no entries
        out = tmp_path / "empty_export.xlsx"
        export_service.export_excel(
            from_date=datetime.date(2026, 4, 11),
            to_date=datetime.date(2026, 4, 12),
            output_path=out,
        )

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # Row 1 = headers, Row 2 = TOTAL summary (no data rows since no workdays)
        assert ws.max_row == 2
        assert ws.cell(row=2, column=1).value == "TOTAL"

    def test_return_value_is_resolved_path(self, db_conn, tmp_path):
        """
        export_excel returns the resolved output_path (TEST-021).
        """
        _, export_service = _setup_services(db_conn)
        out = tmp_path / "path_test.xlsx"
        result = export_service.export_excel(
            from_date=datetime.date(2026, 4, 14),
            to_date=datetime.date(2026, 4, 14),
            output_path=out,
        )
        assert result == out.resolve()
        assert result.exists()

    def test_data_rows_contain_entry_data(self, db_conn, tmp_path):
        """Export with a work entry produces a data row with correct date and type."""
        entry_service, export_service = _setup_services(db_conn)

        entry_service.add_entry(
            date=datetime.date(2026, 4, 14),
            entry_type=EntryType.krank,
        )

        out = tmp_path / "data_test.xlsx"
        export_service.export_excel(
            from_date=datetime.date(2026, 4, 14),
            to_date=datetime.date(2026, 4, 14),
            output_path=out,
        )

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # Row 1 = headers, Row 2 = data, Row 3 = TOTAL
        assert ws.cell(row=2, column=1).value == "2026-04-14"
        assert ws.cell(row=2, column=2).value == "krank"
