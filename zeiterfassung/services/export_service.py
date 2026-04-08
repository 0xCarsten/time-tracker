"""
Excel export service for zeiterfassung.

Constructor injection of EntryService (CRIT-003 fix):
ExportService calls self.entry_service.build_day_results() — it does NOT hold a
direct EntryRepository reference and does NOT reimplement build_day_results.
"""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from zeiterfassung.services.entry_service import EntryService


class ExportService:
    """
    Exports time entries to Excel (.xlsx) format.

    Depends on EntryService for data access (constructor injection, CRIT-003).
    Column layout: Date, Type, Start, End, Pause (h), Delta (h), Running Saldo (h).
    """

    HEADERS = ["Date", "Type", "Start", "End", "Pause (h)", "Delta (h)", "Running Saldo (h)"]

    def __init__(self, entry_service: EntryService) -> None:
        """
        Initialise with an EntryService instance (CRIT-003).

        Parameters:
            entry_service: Used to retrieve day results for export.
        """
        self.entry_service = entry_service

    def export_excel(
        self,
        from_date: datetime.date,
        to_date: datetime.date,
        output_path: Path,
    ) -> Path:
        """
        Export time entries in [from_date, to_date] to an Excel workbook.

        Writes header row (row 1), then one row per DayResult, then a bold
        summary row with totals. Returns the resolved output path.

        Parameters:
            from_date: Start date (inclusive).
            to_date: End date (inclusive).
            output_path: Target .xlsx file path.

        Returns:
            Resolved output_path.
        """
        results = self.entry_service.build_day_results(from_date, to_date)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Zeiterfassung"

        # Header row
        ws.append(self.HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Data rows
        running_saldo = 0
        for result in results:
            running_saldo += result.delta_minutes
            entry = result.entry
            row = [
                result.date.isoformat(),
                entry.entry_type.value if entry else "missing",
                entry.start_time.strftime("%H:%M") if entry and entry.start_time else "",
                entry.end_time.strftime("%H:%M") if entry and entry.end_time else "",
                round(entry.pause_minutes / 60, 2) if entry else 0.0,
                round(result.delta_minutes / 60, 2),
                round(running_saldo / 60, 2),
            ]
            ws.append(row)

        # Summary row
        total_delta = sum(r.delta_minutes for r in results)
        summary = [
            "TOTAL", "", "", "", "",
            round(total_delta / 60, 2),
            round(running_saldo / 60, 2),
        ]
        ws.append(summary)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        wb.save(output_path)
        return output_path.resolve()
